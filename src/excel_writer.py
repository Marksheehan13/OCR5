"""
excel_writer.py

Writes a list of InvoiceExtraction results to a formatted .xlsx workbook.

Upgrades over OCR3's version:
  - More columns: Invoice Date, Supplier, Amount, Currency, Confidence,
    Warnings, Source File, OCR Text, Review Required.
  - Conditional formatting: low-confidence rows, missing values, and
    suspicious amounts (e.g. 0.00, or absurdly large) are highlighted
    so a human reviewer's eye goes straight to what needs checking.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import InvoiceExtraction

HEADER_FILL = PatternFill(start_color="1F2933", end_color="1F2933", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

LOW_CONF_FILL = PatternFill(start_color="FDE2E1", end_color="FDE2E1", fill_type="solid")
MISSING_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
SUSPICIOUS_FILL = PatternFill(start_color="FCE8C7", end_color="FCE8C7", fill_type="solid")

COLUMNS = [
    "Invoice Date",
    "Supplier",
    "Amount",
    "Currency",
    "Confidence",
    "Warnings",
    "Source File",
    "OCR Text",
    "Review Required",
]

SUSPICIOUS_AMOUNT_MAX = 1_000_000.0


def _is_suspicious_amount(amount: str | None) -> bool:
    if amount is None:
        return False
    try:
        value = float(amount)
    except ValueError:
        return True
    return value <= 0 or value > SUSPICIOUS_AMOUNT_MAX


def write_invoices_to_excel(results: list[InvoiceExtraction], output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    for col_idx, header in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_idx, result in enumerate(results, start=2):
        row_values = [
            result.date.value or "",
            result.supplier.value or "",
            result.amount.value or "",
            result.currency,
            result.overall_confidence,
            "; ".join(result.warnings),
            result.source_file,
            result.raw_text[:2000],  # avoid absurdly large cells
            "Yes" if result.needs_review else "No",
        ]
        for col_idx, value in enumerate(row_values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

        # --- Conditional highlighting ---
        row_cells = [ws.cell(row=row_idx, column=c) for c in range(1, len(COLUMNS) + 1)]

        if result.overall_confidence < 70:
            for cell in row_cells:
                cell.fill = LOW_CONF_FILL
        if result.date.value is None or result.supplier.value is None or result.amount.value is None:
            for cell in row_cells:
                cell.fill = MISSING_FILL
        if _is_suspicious_amount(result.amount.value):
            ws.cell(row=row_idx, column=3).fill = SUSPICIOUS_FILL

    widths = [14, 28, 12, 10, 12, 40, 24, 50, 16]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    wb.save(output_path)
