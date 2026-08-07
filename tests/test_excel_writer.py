"""tests/test_excel_writer.py -- same shape as OCR4's, since excel_writer.py's interface is unchanged."""

from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))


def test_write_invoices_to_excel(tmp_path):
    from src.excel_writer import write_invoices_to_excel
    from src.models import FieldResult, InvoiceExtraction

    result = InvoiceExtraction(
        source_file="test.jpg",
        date=FieldResult(value="04/08/2026", confidence=96, reasons=["clearly printed"]),
        supplier=FieldResult(value="Test Vendor Ltd", confidence=90, reasons=["header line"]),
        amount=FieldResult(value="184.50", confidence=87, reasons=["labeled Total Due"]),
        raw_text='{"date": "04/08/2026"}',
    )
    output_path = tmp_path / "out.xlsx"
    write_invoices_to_excel([result], str(output_path))
    assert output_path.exists()

    from openpyxl import load_workbook

    wb = load_workbook(str(output_path))
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "Invoice Date"
    assert ws.cell(row=2, column=1).value == "04/08/2026"
    assert ws.cell(row=2, column=3).value == "184.50"
